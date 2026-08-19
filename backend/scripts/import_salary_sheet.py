"""Import employees + monthly salaries from KAFI JUL-26 salary sheet.

Uses the current-month tab (" Salary Sheet" / JULY-2026). Older tabs and
Cotton & Silk factory sheets are ignored.
"""
from __future__ import annotations

import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.db import get_session_factory  # noqa: E402
from app.models.employees import Department, Employee  # noqa: E402

SKIP_NAME_FRAGMENTS = (
    "grand total",
    "dept wise",
    "prepared by",
    "checked by",
    "approved by",
)
TITLE_WORDS = {"ceo", "it", "ai", "kmp", "asst"}


def _norm_name(value: str) -> str:
    text = (value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _display_name(value: str) -> str:
    raw = re.sub(r"\s+", " ", (value or "").strip())
    raw = re.sub(r"\.(?=\S)", ". ", raw)
    parts = []
    for token in raw.split(" "):
        if not token:
            continue
        core = token.strip(".")
        if core.lower() in TITLE_WORDS or (len(core) <= 2 and core.isalpha()):
            parts.append(token.upper() if token.isalpha() else token[0].upper() + token[1:])
            continue
        parts.append(token[:1].upper() + token[1:].lower() if token else token)
    return " ".join(parts)


def _money(value: object) -> Decimal | None:
    if value is None or isinstance(value, str):
        return None
    try:
        amount = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    except Exception:
        return None
    if amount <= 0:
        return None
    return amount


def parse_july_sheet(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "salary sheet":
            ws = wb[name]
            break
    if ws is None:
        raise SystemExit(f"No 'Salary Sheet' tab in {path}")

    rows: list[dict] = []
    seen: set[str] = set()
    for r in range(5, (ws.max_row or 5) + 1):
        sno = ws.cell(r, 2).value
        name = str(ws.cell(r, 3).value or "").strip()
        designation = str(ws.cell(r, 4).value or "").strip()
        salary = _money(ws.cell(r, 5).value)
        if not isinstance(sno, (int, float)) or not name or not salary:
            continue
        lowered = name.lower()
        if any(frag in lowered for frag in SKIP_NAME_FRAGMENTS):
            continue
        key = _norm_name(name)
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "serial": int(sno),
                "full_name": _display_name(name),
                "match_key": key,
                "designation": re.sub(r"\s+", " ", designation).strip() or "Staff",
                "base_salary": salary,
            }
        )
    return rows


def _ensure_department(db, name: str) -> Department:
    existing = {d.name.strip().lower(): d for d in db.query(Department).all()}
    hit = existing.get(name.strip().lower())
    if hit:
        return hit
    dept = Department(name=name)
    db.add(dept)
    db.flush()
    return dept


def import_rows(db, rows: list[dict]) -> tuple[int, int]:
    employees = db.query(Employee).all()
    by_name: dict[str, Employee] = {}
    for emp in employees:
        key = _norm_name(emp.full_name)
        if key and key not in by_name:
            by_name[key] = emp

    created = 0
    updated = 0
    used_codes = {e.employee_code for e in employees}

    for row in rows:
        emp = by_name.get(row["match_key"])
        dept = _ensure_department(db, row["designation"])
        if emp is None:
            code = f"KC{row['serial']:03d}"
            if code in used_codes:
                n = 1
                while f"{code}-{n}" in used_codes:
                    n += 1
                code = f"{code}-{n}"
            emp = Employee(
                employee_code=code,
                full_name=row["full_name"],
                department_id=dept.id,
                role_title=row["designation"],
                employment_type="full_time",
                status="active",
                base_salary=row["base_salary"],
            )
            db.add(emp)
            db.flush()
            used_codes.add(code)
            by_name[row["match_key"]] = emp
            created += 1
            print(f"  created {code:8}  {emp.full_name:28}  {row['designation']:22}  {row['base_salary']}")
            continue

        emp.full_name = row["full_name"]
        emp.department_id = dept.id
        emp.role_title = row["designation"]
        emp.base_salary = row["base_salary"]
        if emp.status == "terminated":
            emp.status = "active"
            emp.date_exited = None
        updated += 1
        print(
            f"  updated {emp.employee_code:8}  {emp.full_name:28}  "
            f"{row['designation']:22}  {row['base_salary']}"
        )
    return created, updated


def main() -> None:
    default = Path(r"c:\Users\User\Downloads\SALARY SHEET JUL-26.xlsx")
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not path.is_file():
        raise SystemExit(f"Excel file not found: {path}")

    rows = parse_july_sheet(path)
    if not rows:
        raise SystemExit("No employee rows found on the July 2026 salary sheet")

    print(f"Parsed {len(rows)} employees from {path.name}")
    SessionLocal = get_session_factory()
    db = SessionLocal()
    try:
        created, updated = import_rows(db, rows)
        db.commit()
        print(f"Done. created={created} updated={updated}")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
