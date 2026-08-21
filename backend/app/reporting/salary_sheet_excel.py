"""Salary sheet Excel matching KAFI COMMODITIES empty format (visible columns)."""
from __future__ import annotations

from calendar import month_name
from io import BytesIO
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.payroll import PayrollComputeResult

COMPANY_NAME = "KAFI COMMODITIES (PVT) LTD"

_HEADERS = [
    ("S.No#", "Employee's Detail"),
    ("Name", "Employee's Detail"),
    ("Designation", "Employee's Detail"),
    ("Salary", "Salary"),
    ("Per Day Salary", "Per Day"),
    ("P", "Attendance"),
    ("A", "Attendance"),
    ("OT", "Attendance"),
    ("Half Day", "Half Day"),
    ("Amount", "Allowance"),
    ("Gross Salary", "Gross"),
    ("Late Coming", "Late Coming"),
    ("Late Deduction Amount", "Late Deduction Amount"),
    ("Half Deduction", "Half Deduction"),
    ("Loan Deduction Amount", "Loan Deduction Amount"),
    ("Advance", "Advance"),
    ("Tax/Other Deduction", "Tax/Other Deduction"),
    ("Net Payable", "Net Payable"),
    ("Mode of Payment", "Mode of Payment"),
    ("Remarks", "Remarks"),
]

_COL_WIDTHS = [8, 22, 22, 12, 12, 8, 8, 8, 10, 12, 14, 12, 16, 14, 16, 12, 16, 14, 14, 28]


def _money(n) -> float:
    return float(n or 0)


def build_salary_sheet_workbook(result: PayrollComputeResult) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Sheet"

    thin = Border(
        left=Side(style="thin", color="B8C1CD"),
        right=Side(style="thin", color="B8C1CD"),
        top=Side(style="thin", color="B8C1CD"),
        bottom=Side(style="thin", color="B8C1CD"),
    )
    title_fill = PatternFill("solid", fgColor="2B4C7E")
    title_font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    sub_fill = PatternFill("solid", fgColor="E5EBF3")
    sub_font = Font(name="Calibri", bold=True, color="101828", size=12)
    head_fill = PatternFill("solid", fgColor="EEF1F5")
    head_font = Font(name="Calibri", bold=True, color="4B5567", size=9)
    money_font = Font(name="Consolas", size=10)
    total_fill = PatternFill("solid", fgColor="E5EBF3")

    last_col = len(_HEADERS)
    last_letter = get_column_letter(last_col)
    month_label = f"{month_name[result.period_month].upper()}-{result.period_year}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c1 = ws.cell(1, 1, result.company_name or COMPANY_NAME)
    c1.font = title_font
    c1.fill = title_fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c2 = ws.cell(2, 1, f"Salary Sheet For The Month Of {month_label}")
    c2.font = sub_font
    c2.fill = sub_fill
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Group row + sub-header row (Excel rows 3–4)
    group_ranges: list[tuple[int, int, str]] = []
    start = 1
    current = _HEADERS[0][1]
    for i, (_, group) in enumerate(_HEADERS, start=1):
        if group != current:
            group_ranges.append((start, i - 1, current))
            start = i
            current = group
    group_ranges.append((start, last_col, current))

    for a, b, label in group_ranges:
        if a != b:
            ws.merge_cells(start_row=3, start_column=a, end_row=3, end_column=b)
        cell = ws.cell(3, a, label)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        for col in range(a, b + 1):
            ws.cell(3, col).fill = head_fill
            ws.cell(3, col).border = thin

    for i, (sub, _) in enumerate(_HEADERS, start=1):
        cell = ws.cell(4, i, sub)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    money_cols = {4, 5, 10, 11, 13, 14, 15, 16, 17, 18}

    for idx, emp in enumerate(result.employees, start=1):
        r = idx + 4
        values = [
            idx,
            emp.full_name,
            emp.role_title,
            _money(emp.base_salary),
            _money(emp.per_day_rate),
            emp.days_present,
            emp.absents_after_leave,
            emp.overtime_bonus_days,
            emp.days_half_day,
            _money(emp.allowance_amount),
            _money(emp.gross_salary),
            emp.days_late,
            _money(emp.late_deduction_amount),
            _money(emp.half_day_deduction),
            _money(emp.loan_deduction_amount),
            _money(emp.advance_amount),
            _money(emp.monthly_tax),
            _money(emp.net_payable),
            emp.payment_mode or "IBFT",
            emp.remarks or "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(r, col, val)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=col in (2, 3, 20))
            if col in money_cols:
                cell.number_format = "#,##0.00"
                cell.font = money_font
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in (1, 6, 7, 8, 9, 12):
                cell.font = money_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row = 5 + len(result.employees)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    tcell = ws.cell(total_row, 1, "Grand Total")
    tcell.font = Font(name="Calibri", bold=True, size=10)
    tcell.fill = total_fill
    tcell.border = thin
    tcell.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(2, last_col + 1):
        ws.cell(total_row, col).fill = total_fill
        ws.cell(total_row, col).border = thin
    for col in (4, 5, 10, 11, 13, 14, 15, 16, 17, 18):
        letter = get_column_letter(col)
        cell = ws.cell(total_row, col, f"=SUM({letter}5:{letter}{total_row - 1})" if result.employees else 0)
        cell.number_format = "#,##0.00"
        cell.font = Font(name="Consolas", bold=True, size=10)
        cell.fill = total_fill
        cell.border = thin

    # Payment mode summary sheet (report figures — not shown as a UI banner)
    from collections import Counter

    from app.services.payroll_service import _normalize_payment_mode

    counts: Counter[str] = Counter()
    net_by_mode: dict[str, float] = {"IBFT": 0.0, "Cash": 0.0, "Cheque": 0.0}
    for emp in result.employees:
        mode = _normalize_payment_mode(emp.payment_mode)
        counts[mode] += 1
        net_by_mode[mode] = net_by_mode.get(mode, 0.0) + _money(emp.net_payable)

    summary = wb.create_sheet("Payment Summary")
    summary["A1"] = "Payment mode summary"
    summary["A1"].font = Font(name="Calibri", bold=True, size=12, color="101828")
    summary["A2"] = f"Salary sheet — {month_label}"
    summary["A3"] = "Mode"
    summary["B3"] = "Employees"
    summary["C3"] = "Net payable"
    for col in ("A", "B", "C"):
        summary[f"{col}3"].font = head_font
        summary[f"{col}3"].fill = head_fill
    for i, mode in enumerate(("IBFT", "Cash", "Cheque"), start=4):
        summary.cell(i, 1, mode)
        summary.cell(i, 2, int(counts.get(mode, 0)))
        cell = summary.cell(i, 3, net_by_mode.get(mode, 0.0))
        cell.number_format = "#,##0.00"
        cell.font = money_font
    summary.cell(7, 1, "Total employees")
    summary.cell(7, 2, len(result.employees))
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 16

    sign_row = total_row + 2
    ws.cell(sign_row, 2, "Prepared By")
    ws.cell(sign_row, 8, "Checked By")
    ws.cell(sign_row, 14, "Approved By")
    for col in (2, 8, 14):
        ws.cell(sign_row, col).font = Font(name="Calibri", italic=True, color="4B5567", size=10)

    for i, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_letter}{max(total_row - 1, 4)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:4"
    ws.sheet_view.showGridLines = False
    return wb


def salary_sheet_xlsx_bytes(result: PayrollComputeResult) -> bytes:
    wb = build_salary_sheet_workbook(result)
    buf: BinaryIO = BytesIO()
    wb.save(buf)
    return buf.getvalue()
