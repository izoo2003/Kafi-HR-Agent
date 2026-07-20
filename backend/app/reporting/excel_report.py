"""Generates HR-facing Excel workbooks from ranked, scored applications —
formatted (colors, banded verdicts, wrapped summaries) so it's presentable
straight out of the pipeline, no manual cleanup needed.

Two report types:
- Per-position report: overview ranking table (with decision summary) +
  detailed candidate profiles.
- Master report: position summary sheet + all-candidates sheet with a
  why-hire / why-reject description for every person.
"""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.config import settings
from app.db.models import Application
from app.ranking.ranker import get_ranked_applications, list_positions

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)

VERDICT_FILLS = {
    "STRONG HIRE": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "RECOMMEND": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
    "CONDITIONAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NOT RECOMMENDED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

POSITIVE_VERDICTS = {"STRONG HIRE", "RECOMMEND", "CONDITIONAL"}


def _style_header_row(ws: Worksheet, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_columns(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _title_banner(ws: Worksheet, text: str, num_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28


def _decision_label(verdict: str | None) -> str:
    if verdict in {"STRONG HIRE", "RECOMMEND"}:
        return "SELECT"
    if verdict == "CONDITIONAL":
        return "CONDITIONAL"
    return "REJECT"


def _decision_summary(app_: Application) -> str:
    """Builds a clear hire/reject description for Excel. Prefers the LLM
    hiring_summary; adds an explicit SELECT/REJECT prefix for HR scanning."""
    decision = _decision_label(app_.verdict)
    body = (app_.hiring_summary or "").strip()
    if not body:
        if decision == "REJECT":
            body = (
                "Insufficient evidence on the CV for this role. "
                "Not recommended for hire at this time."
            )
        elif decision == "CONDITIONAL":
            body = (
                "Emerging or partial fit for this role. "
                "Consider only if junior capacity or niche needs apply."
            )
        else:
            body = (
                "Meets role expectations based on available CV evidence. "
                "Recommended to advance to interview."
            )

    if decision == "SELECT":
        prefix = "Why select:"
    elif decision == "CONDITIONAL":
        prefix = "Why conditional:"
    else:
        prefix = "Why reject:"

    if body.lower().startswith(("why select", "why reject", "why conditional")):
        return body
    return f"{prefix} {body}"


def _apply_row_fill(ws: Worksheet, row: int, num_cols: int, verdict: str | None) -> None:
    fill = VERDICT_FILLS.get(verdict or "")
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def _prune_older_duplicates(kept: Path, glob_pattern: str) -> list[Path]:
    """Keep only `kept`; delete older reports matching the same series pattern.

    Called after every new report save so regenerations never leave duplicate
    Master or per-position Excel files in the reports folder.
    """
    reports_dir = settings.reports_output_dir
    if not reports_dir.exists():
        return []

    kept_resolved = kept.resolve()
    deleted: list[Path] = []
    for path in reports_dir.glob(glob_pattern):
        if not path.is_file():
            continue
        if path.resolve() == kept_resolved:
            continue
        path.unlink(missing_ok=True)
        deleted.append(path)
    return deleted


def generate_position_report(session: Session, position: str) -> Path:
    applications = get_ranked_applications(session, position)

    wb = Workbook()
    _write_overview_sheet(wb.active, position, applications)
    _write_detail_sheet(wb.create_sheet("Detailed Profiles"), applications)

    settings.reports_output_dir.mkdir(parents=True, exist_ok=True)
    safe_position = "".join(c if c.isalnum() else "_" for c in position)
    filename = f"CV_Ranking_{safe_position}_{dt.date.today().isoformat()}.xlsx"
    dest = settings.reports_output_dir / filename
    wb.save(dest)
    _prune_older_duplicates(dest, f"CV_Ranking_{safe_position}_*.xlsx")
    return dest


def _write_overview_sheet(ws: Worksheet, position: str, applications: list[Application]) -> None:
    ws.title = "Overview"
    headers = [
        "Rank",
        "Candidate",
        "Position",
        "Score",
        "Verdict",
        "Decision",
        "Why Select / Why Reject",
        "Email",
        "Phone",
        "Source",
    ]
    _title_banner(ws, f"CV Ranking Report — {position}", len(headers))

    ws.append([])  # spacer row
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for app_ in applications:
        ws.cell(row=row, column=1, value=app_.rank_in_position)
        ws.cell(row=row, column=2, value=app_.candidate.full_name)
        ws.cell(row=row, column=3, value=app_.position_applied)
        ws.cell(row=row, column=4, value=app_.score)
        ws.cell(row=row, column=5, value=app_.verdict)
        ws.cell(row=row, column=6, value=_decision_label(app_.verdict))
        ws.cell(row=row, column=7, value=_decision_summary(app_))
        ws.cell(row=row, column=8, value=app_.candidate.email)
        ws.cell(row=row, column=9, value=app_.candidate.phone or "")
        ws.cell(row=row, column=10, value=app_.source.value)

        _apply_row_fill(ws, row, len(headers), app_.verdict)
        ws.row_dimensions[row].height = 75
        row += 1

    if applications:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = f"A{header_row + 1}"
    _autosize_columns(ws, [8, 24, 20, 8, 16, 12, 55, 28, 14, 12])


def _write_detail_sheet(ws: Worksheet, applications: list[Application]) -> None:
    headers = [
        "Rank",
        "Candidate",
        "Score",
        "Verdict",
        "Decision",
        "Education",
        "Experience",
        "Key Strengths",
        "Why Select / Why Reject",
        "Location",
    ]
    header_row = 1
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for app_ in applications:
        strengths = json.loads(app_.key_strengths_json) if app_.key_strengths_json else []
        strengths_text = "\n".join(f"• {s}" for s in strengths)

        ws.cell(row=row, column=1, value=app_.rank_in_position)
        ws.cell(row=row, column=2, value=app_.candidate.full_name)
        ws.cell(row=row, column=3, value=app_.score)
        ws.cell(row=row, column=4, value=app_.verdict)
        ws.cell(row=row, column=5, value=_decision_label(app_.verdict))
        ws.cell(row=row, column=6, value=app_.education_summary or "")
        ws.cell(row=row, column=7, value=app_.experience_summary or "")
        ws.cell(row=row, column=8, value=strengths_text)
        ws.cell(row=row, column=9, value=_decision_summary(app_))
        ws.cell(row=row, column=10, value=app_.candidate.location or "")

        _apply_row_fill(ws, row, len(headers), app_.verdict)
        ws.row_dimensions[row].height = 110
        row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize_columns(ws, [6, 24, 8, 16, 12, 30, 30, 36, 55, 16])


def generate_master_report(session: Session) -> Path:
    positions = list_positions(session)

    wb = Workbook()
    # All Candidates first so the why-select / why-reject text is visible
    # immediately when the downloaded file opens (not buried on sheet 2).
    _write_master_all_candidates_sheet(wb.active, positions, session)
    _write_master_summary_sheet(wb.create_sheet("By Position"), positions, session)

    settings.reports_output_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = settings.reports_output_dir / f"CV_Ranking_Master_{stamp}.xlsx"
    wb.save(dest)
    _prune_older_duplicates(dest, "CV_Ranking_Master_*.xlsx")
    return dest


def _write_master_summary_sheet(
    ws: Worksheet, positions: list[str], session: Session
) -> None:
    ws.title = "By Position"
    headers = [
        "Position",
        "Candidates Scored",
        "Top Candidate",
        "Top Score",
        "Top Verdict",
        "Top Decision Summary",
    ]
    _title_banner(ws, "HR & Admin Agent — CV Ranking Master Summary", len(headers))
    ws.append([])
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for position in positions:
        applications = get_ranked_applications(session, position)
        top = applications[0] if applications else None
        ws.cell(row=row, column=1, value=position)
        ws.cell(row=row, column=2, value=len(applications))
        ws.cell(row=row, column=3, value=top.candidate.full_name if top else "")
        ws.cell(row=row, column=4, value=top.score if top else "")
        ws.cell(row=row, column=5, value=top.verdict if top else "")
        ws.cell(row=row, column=6, value=_decision_summary(top) if top else "")
        if top:
            _apply_row_fill(ws, row, len(headers), top.verdict)
            ws.row_dimensions[row].height = 60
        row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize_columns(ws, [24, 16, 24, 10, 16, 55])


def _write_master_all_candidates_sheet(
    ws: Worksheet, positions: list[str], session: Session
) -> None:
    ws.title = "All Candidates"
    headers = [
        "Position",
        "Rank",
        "Candidate",
        "Score",
        "Verdict",
        "Decision",
        "Why Select / Why Reject",
        "Email",
        "Source",
    ]
    _title_banner(
        ws,
        "HR & Admin Agent — All Candidates (Why Select / Why Reject)",
        len(headers),
    )
    ws.append([])
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for position in positions:
        for app_ in get_ranked_applications(session, position):
            ws.cell(row=row, column=1, value=app_.position_applied)
            ws.cell(row=row, column=2, value=app_.rank_in_position)
            ws.cell(row=row, column=3, value=app_.candidate.full_name)
            ws.cell(row=row, column=4, value=app_.score)
            ws.cell(row=row, column=5, value=app_.verdict)
            ws.cell(row=row, column=6, value=_decision_label(app_.verdict))
            ws.cell(row=row, column=7, value=_decision_summary(app_))
            ws.cell(row=row, column=8, value=app_.candidate.email)
            ws.cell(row=row, column=9, value=app_.source.value)
            _apply_row_fill(ws, row, len(headers), app_.verdict)
            ws.row_dimensions[row].height = 80
            row += 1

    if row > header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = f"A{header_row + 1}"
    _autosize_columns(ws, [22, 6, 24, 8, 16, 12, 55, 28, 12])
