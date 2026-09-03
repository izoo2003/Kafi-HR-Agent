"""Office attendance Excel/CSV period analysis.

Canonical upload format (WebHR-style export):
- Optional title row, then header row with: Employee ID, First Name, Department,
  Date, Weekday, First Punch, Last Punch, Total Time
- One row per person per day that has a punch (sparse — missing working days = absent)
- Identity for the report = Excel display name (First Name / First & Last column)
- First Punch only drives late / half-day / presence; Last Punch ignored for rules
- Period = full calendar month(s) inferred from dates in the file
- Excel L / WAVE flags ignored (own calculations are source of truth)

Policy (Kafi office):
- Working week: Mon–Sat; Sunday always official off (never counts as absent)
- Late and half-day still count as present (they showed up); tracked separately
- One Saturday off per month: HR selects the date (default Recommended = 2nd Saturday)
  or picks a specific Saturday
- Extra holidays can be entered manually on upload (gazetted / extra company days)
- Presence on those off days (or Sunday) → +1 OT, not present
- 3 lates = 1 late absent (salary deduction), not a regular absent (no-show only)
- Tenure ≥ 6 months → 1 leave allowance / month (only when name matches an Employee)
- Salary math uses fixed 30-day month (base from matched Employee when available)
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.exceptions import ValidationFailed
from app.core.gemini_client import generate_content_with_fallback
from app.models.attendance import AttendanceRecord
from app.models.employees import Department, Employee
from app.models.system import SystemConfig
from app.schemas.attendance import (
    AttendanceEmployeesFromExcelCreate,
    AttendanceEmployeesFromExcelResult,
    AttendancePeriodReport,
    DayClassification,
    ImportErrorRow,
    LateEvent,
    PeriodEmployeeReport,
    PeriodDayEntry,
    PayrollPeriodSaved,
    UnmatchedAttendancePerson,
)
from app.schemas.payroll import PayrollSheetAdjustmentInput, PayrollSheetAdjustmentsSave
from app.schemas.common import AuthContext
from app.services import audit_service
from app.services.attendance_service import (
    _company_tz,
    _holiday_dates,
    _parse_dt,
    build_employee_indexes,
    match_employee,
    merge_holiday_dates,
)

logger = logging.getLogger(__name__)

LATE_AFTER = time(9, 40)
HALF_DAY_AFTER = time(11, 30)
MAJORITY_ABSENT = 0.90
LATES_PER_OFF = 3
MONTH_DAYS = 30
LEAVE_AFTER_MONTHS = 6
MONTHLY_LEAVE = 1

NAME_ALIASES = (
    "name",
    "employee_name",
    "full_name",
    "emp_name",
    "employee",
    "first name",
    "first & last",
    "first and last",
)
CODE_ALIASES = (
    "employee_code",
    "emp_code",
    "code",
    "employee id",
    "emp_id",
    "emp id",
    "staff id",
    "id no",
)
LAST_NAME_ALIASES = ("last name", "surname", "last_name", "family name")
EMAIL_ALIASES = ("email", "work_email", "email address", "e-mail")
DATE_ALIASES = ("date", "attendance_date", "day", "punch_date")
IN_ALIASES = (
    "check_in",
    "checkin",
    "time_in",
    "in",
    "punch_in",
    "clock_in",
    "time",
    "first punch",
    "firstpunch",
)
OUT_ALIASES = (
    "check_out",
    "checkout",
    "time_out",
    "out",
    "punch_out",
    "clock_out",
    "last punch",
    "lastpunch",
)

HEADER_MARKERS = frozenset(
    {
        "employee id",
        "first name",
        "date",
        "first punch",
        "last punch",
        "weekday",
        "department",
    }
)


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower().replace("_", " "))


def _find_col(field_map: dict[str, str], aliases: tuple[str, ...]) -> str | None:
    for a in aliases:
        key = _norm_header(a)
        if key in field_map:
            return field_map[key]
    return None


def _normalize_excel_id(raw: str) -> str:
    s = (raw or "").strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


def _norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    return str(val).strip()


def _is_header_row(values: list[str]) -> bool:
    norms = {_norm_header(v) for v in values if v}
    hits = len(norms & HEADER_MARKERS)
    return hits >= 3 or "first punch" in norms or ("date" in norms and "first name" in norms)


def _row_to_item(headers: list[str], row: tuple | list) -> dict[str, str]:
    item: dict[str, str] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        val = row[i] if i < len(row) else None
        item[h] = _cell_str(val)
    return item


def _parse_excel_or_csv(content: bytes, filename: str) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationFailed("Excel file is empty")

        header_idx: int | None = None
        headers: list[str] = []
        for i, row in enumerate(rows[:15]):
            vals = ["" if c is None else str(c).strip() for c in row]
            if _is_header_row(vals):
                header_idx = i
                headers = vals
                break
        if header_idx is None:
            # Fallback: first non-empty row as headers (legacy templates)
            for i, row in enumerate(rows):
                if row and any(c is not None and str(c).strip() for c in row):
                    header_idx = i
                    headers = ["" if c is None else str(c).strip() for c in row]
                    break
        if header_idx is None:
            raise ValidationFailed("Could not find header row (need Date + First Punch / First Name)")

        out: list[dict[str, str]] = []
        for row in rows[header_idx + 1 :]:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            vals = ["" if c is None else str(c).strip() for c in row]
            if _is_header_row(vals):
                continue
            item = _row_to_item(headers, row)
            # Skip junk / section labels
            name_guess = ""
            for k, v in item.items():
                if _norm_header(k) in ("first name", "name", "full name", "employee"):
                    name_guess = v
                    break
            if _norm_header(name_guess) in HEADER_MARKERS or name_guess.upper() in (
                "SUNDAY",
                "HOLIDAY",
                "SATURDAY",
            ):
                continue
            out.append(item)
        return out

    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValidationFailed("CSV has no header row")
    return [{k: ("" if v is None else str(v).strip()) for k, v in row.items()} for row in reader]


def _parse_date_value(raw: str, tz: ZoneInfo) -> date:
    raw = (raw or "").strip()
    if not raw:
        raise ValueError("Empty date")
    try:
        if re.fullmatch(r"\d+(\.\d+)?", raw):
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(float(raw))
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
    except Exception:
        pass
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(raw[:20], fmt).date()
        except ValueError:
            continue
    dt = _parse_dt(raw, date.today(), tz)
    if dt is None:
        raise ValueError(f"Invalid date: {raw}")
    return dt.astimezone(tz).date()


def _tenure_months(joined: date | None, as_of: date) -> int:
    if joined is None or joined > as_of:
        return 0
    months = (as_of.year - joined.year) * 12 + (as_of.month - joined.month)
    if as_of.day < joined.day:
        months -= 1
    return max(0, months)


def _office_policy(db: Session) -> dict:
    row = db.query(SystemConfig).filter(SystemConfig.key == "attendance.office_policy").one_or_none()
    defaults = {
        "late_after": "09:40",
        "half_day_after": "11:30",
        "majority_absent_threshold": MAJORITY_ABSENT,
        "lates_per_off": LATES_PER_OFF,
        "month_days": MONTH_DAYS,
        "leave_after_months": LEAVE_AFTER_MONTHS,
        "monthly_leave_allowance": MONTHLY_LEAVE,
    }
    if row and isinstance(row.value, dict):
        return {**defaults, **row.value}
    return defaults


def _time_from_policy(value: str, fallback: time) -> time:
    try:
        parts = str(value).split(":")
        return time(int(parts[0]), int(parts[1]))
    except Exception:
        return fallback


def _calendar_period(dates: list[date]) -> tuple[date, date]:
    """Option A: full calendar month(s) covering dates found in the file."""
    first = min(dates)
    last = max(dates)
    start = date(first.year, first.month, 1)
    end = date(last.year, last.month, monthrange(last.year, last.month)[1])
    return start, end


def _nth_weekday_of_month(year: int, month: int, weekday: int, n: int) -> date | None:
    """weekday: Mon=0 … Sun=6. n is 1-based (2 = second occurrence)."""
    first = date(year, month, 1)
    # Days until the first desired weekday
    offset = (weekday - first.weekday()) % 7
    day = 1 + offset + (n - 1) * 7
    last_day = monthrange(year, month)[1]
    if day > last_day:
        return None
    return date(year, month, day)


def _second_saturdays(period_start: date, period_end: date) -> set[date]:
    """Second Saturday of each calendar month that overlaps the period."""
    out: set[date] = set()
    y, m = period_start.year, period_start.month
    while True:
        cursor = date(y, m, 1)
        if cursor > period_end:
            break
        sat = _nth_weekday_of_month(y, m, weekday=5, n=2)
        if sat is not None and period_start <= sat <= period_end:
            out.add(sat)
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
    return out


def _saturdays_in_period(period_start: date, period_end: date) -> list[date]:
    out: list[date] = []
    d = period_start
    while d <= period_end:
        if d.weekday() == 5:
            out.append(d)
        d += timedelta(days=1)
    return out


def _absent_rate_for_day(
    day: date,
    roster_keys: list[str],
    punches: dict[str, dict[date, dict[str, datetime | None]]],
) -> float:
    present_n = sum(1 for k in roster_keys if punches.get(k, {}).get(day, {}).get("check_in"))
    total_n = max(len(roster_keys), 1)
    return 1.0 - (present_n / total_n)


def _detect_saturday_off_with_ai(
    saturdays: list[date],
    rates: dict[date, float],
    majority: float,
) -> set[date]:
    """When HR picks Don't know: ask Gemini which Saturday is company off; else heuristic."""
    if not saturdays:
        return set()

    # Prefer Saturdays that look empty (≥ majority absent)
    candidates = [d for d in saturdays if rates.get(d, 0) >= majority]
    if not candidates:
        # Still pick the emptiest Saturday so we don't invent absences for a full workday
        best = max(saturdays, key=lambda d: rates.get(d, 0.0))
        if rates.get(best, 0.0) >= 0.5:
            candidates = [best]
        else:
            return set()

    settings = get_settings()
    api_keys = settings.resolved_gemini_api_keys()
    if not api_keys:
        # Heuristic: one Saturday per month — the emptiest that meets threshold
        by_month: dict[tuple[int, int], date] = {}
        for d in sorted(candidates, key=lambda x: (-rates.get(x, 0), x)):
            key = (d.year, d.month)
            if key not in by_month:
                by_month[key] = d
        return set(by_month.values())

    payload = [
        {"date": d.isoformat(), "weekday": d.strftime("%A"), "absent_rate": round(rates[d], 3)}
        for d in saturdays
    ]
    prompt = f"""You are helping HR pick the company's monthly Saturday off from attendance punches.

Policy: usually the second Saturday is off, but HR did not know this month's date.
Pick at most one Saturday per calendar month that is most likely the company off day
(high absent_rate means almost nobody punched — strong off-day signal).
Threshold hint: absent_rate >= {majority} is typical for an off day.

Saturdays in this file (JSON):
{json.dumps(payload, indent=2)}

Respond with STRICT JSON only, no markdown:
{{"saturday_off_dates": ["YYYY-MM-DD", ...]}}
If none look like an off day, return an empty list.
"""
    try:
        response = generate_content_with_fallback(
            api_keys=api_keys,
            models=settings.resolved_gemini_models(),
            prompt=prompt,
            pool_id="attendance",
        )
        text = (getattr(response, "text", "") or "").strip()
        fence = re.search(r"\{.*\}", text, re.DOTALL)
        if fence:
            text = fence.group(0)
        data = json.loads(text)
        picked: set[date] = set()
        valid = {d.isoformat(): d for d in saturdays}
        for raw in data.get("saturday_off_dates") or []:
            key = str(raw).strip()[:10]
            if key in valid:
                picked.add(valid[key])
        if picked:
            return picked
    except Exception as exc:  # noqa: BLE001
        logger.warning("Saturday-off AI detection failed, using heuristic: %s", exc)

    by_month: dict[tuple[int, int], date] = {}
    for d in sorted(candidates, key=lambda x: (-rates.get(x, 0), x)):
        key = (d.year, d.month)
        if key not in by_month:
            by_month[key] = d
    return set(by_month.values())


def _require_single_month(dates: list[date]) -> tuple[date, date]:
    """Attendance files are always one calendar month. Period = 1st → last day of that month."""
    period_start, period_end = _calendar_period(dates)
    if (period_start.year, period_start.month) != (period_end.year, period_end.month):
        raise ValidationFailed(
            "This attendance file covers more than one month. Upload a single month only "
            f"(found {period_start.strftime('%B %Y')} to {period_end.strftime('%B %Y')})."
        )
    return period_start, period_end


def _tax_year_for_period(db: Session, as_of: date):
    from app.models.payroll import TaxYear
    from app.services.tax_service import ensure_default_tax_year

    ensure_default_tax_year(db)
    covering = (
        db.query(TaxYear)
        .filter(TaxYear.start_date <= as_of, TaxYear.end_date >= as_of)
        .order_by(TaxYear.is_active.desc(), TaxYear.id.desc())
        .first()
    )
    if covering is not None:
        return covering
    return (
        db.query(TaxYear).filter(TaxYear.is_active.is_(True)).order_by(TaxYear.id.desc()).first()
        or db.query(TaxYear).order_by(TaxYear.id.desc()).first()
    )


def _persist_payroll_for_import(
    db: Session,
    auth: AuthContext,
    *,
    period_start: date,
    period_end: date,
    matched_employee_ids: set[int],
) -> tuple[list[PayrollPeriodSaved], str | None]:
    """Overwrite the salary sheet for the single month in the attendance file."""
    from app.services.payroll_compute import compute_payroll_for_month
    from app.services.payroll_service import save_sheet_adjustments

    if not matched_employee_ids:
        return [], "No matched employees — payroll was not saved. Add them as employees and re-upload professionally."

    tax_year = _tax_year_for_period(db, period_end)
    if tax_year is None:
        return [], "Attendance was saved, but payroll was not — no tax year is configured."

    period_month, period_year = period_start.month, period_start.year
    saved_periods: list[PayrollPeriodSaved] = []
    computed = compute_payroll_for_month(
        db,
        period_month=period_month,
        period_year=period_year,
        tax_year_id=tax_year.id,
        ignore_attendance_overrides=True,
    )
    items = [
        PayrollSheetAdjustmentInput(
            employee_id=row.employee_id,
            base_salary=row.base_salary,
            days_present=row.days_present,
            days_absent=row.absents_after_leave,
            days_late=row.days_late,
            days_half_day=row.days_half_day,
            leave_used=row.leave_used,
            overtime_bonus_days=row.overtime_bonus_days,
            allowance_amount=row.allowance_amount,
            bonus_amount=row.bonus_amount,
            loan_deduction_amount=row.loan_deduction_amount,
            advance_amount=row.advance_amount,
            payment_mode=row.payment_mode,
            remarks=row.remarks,
            monthly_tax_override=row.monthly_tax if row.tax_manual else None,
            excluded=False,
        )
        for row in computed.employees
        if row.employee_id in matched_employee_ids
    ]
    if items:
        count = save_sheet_adjustments(
            db,
            auth,
            PayrollSheetAdjustmentsSave(
                period_month=period_month,
                period_year=period_year,
                items=items,
            ),
        )
        saved_periods.append(
            PayrollPeriodSaved(
                period_month=period_month,
                period_year=period_year,
                employees_saved=count,
            )
        )
    if not saved_periods:
        return [], "Attendance was saved, but no payroll rows were written for this file's month."
    return saved_periods, None


def analyze_period_file(
    db: Session,
    auth: AuthContext,
    content: bytes,
    filename: str,
    *,
    persist: bool = True,
    saturday_off_mode: str = "second_saturday",
    saturday_off_date: date | None = None,
    extra_holiday_dates: list[date] | None = None,
    import_mode: str | None = None,
) -> AttendancePeriodReport:
    policy = _office_policy(db)
    late_after = _time_from_policy(policy["late_after"], LATE_AFTER)
    half_after = _time_from_policy(policy["half_day_after"], HALF_DAY_AFTER)
    majority = float(policy["majority_absent_threshold"])
    lates_per_off = int(policy["lates_per_off"])
    month_days = int(policy["month_days"])
    leave_after_months = int(policy["leave_after_months"])
    monthly_leave = int(policy["monthly_leave_allowance"])

    tz = _company_tz(db)
    extra_holidays = list(extra_holiday_dates or [])
    raw_mode = (import_mode or "").strip().lower()
    if raw_mode in ("testing", "professional"):
        mode_norm = raw_mode
    else:
        mode_norm = "professional" if persist else "testing"
    persist = mode_norm == "professional"
    # Testing: extra holidays apply to this analysis only. Professional: merge into company holidays.
    if persist and extra_holidays:
        configured_holidays = merge_holiday_dates(db, extra_holidays)
    else:
        configured_holidays = _holiday_dates(db) | set(extra_holidays)
    rows = _parse_excel_or_csv(content, filename)
    if not rows:
        raise ValidationFailed("No data rows found in file")

    headers = list(rows[0].keys())
    field_map = {_norm_header(h): h for h in headers}
    name_col = _find_col(field_map, NAME_ALIASES)
    last_col = _find_col(field_map, LAST_NAME_ALIASES)
    code_col = _find_col(field_map, CODE_ALIASES)
    email_col = _find_col(field_map, EMAIL_ALIASES)
    date_col = _find_col(field_map, DATE_ALIASES)
    in_col = _find_col(field_map, IN_ALIASES)
    out_col = _find_col(field_map, OUT_ALIASES)

    if date_col is None:
        raise ValidationFailed("Missing date column (e.g. Date)")
    if in_col is None:
        raise ValidationFailed("Missing First Punch / check-in column")
    if name_col is None and code_col is None:
        raise ValidationFailed("Missing First Name / name or Employee ID column")

    indexes = build_employee_indexes(db)
    errors: list[ImportErrorRow] = []
    # person_key -> date -> first punch (check_out stored only for persist)
    punches: dict[str, dict[date, dict[str, datetime | None]]] = defaultdict(dict)
    display_names: dict[str, str] = {}
    excel_ids: dict[str, str] = {}
    linked_employee: dict[str, Employee] = {}

    for idx, raw in enumerate(rows, start=2):
        try:
            first = (raw.get(name_col) or "").strip() if name_col else ""
            last = (raw.get(last_col) or "").strip() if last_col else ""
            display = " ".join(p for p in (first, last) if p)
            code = (raw.get(code_col) or "").strip() if code_col else ""
            email = (raw.get(email_col) or "").strip() if email_col else ""
            if not display and not code:
                raise ValueError("Missing name and employee id")

            emp = match_employee(code=code, name=display or None, email=email or None, indexes=indexes)
            key = f"id:{emp.id}" if emp is not None else f"name:{_norm_name(display or code)}"
            display_names.setdefault(key, emp.full_name if emp is not None else display or code)
            excel_id = _normalize_excel_id(code)
            if excel_id and key not in excel_ids:
                excel_ids[key] = excel_id
            if emp is not None:
                linked_employee[key] = emp

            on_date = _parse_date_value(raw.get(date_col) or "", tz)
            cin_raw = (raw.get(in_col) or "").strip()
            if not cin_raw:
                # No first punch → ignore row (absence inferred from missing days)
                continue
            cin = _parse_dt(cin_raw, on_date, tz)
            if cin is None:
                raise ValueError(f"Invalid First Punch: {cin_raw}")
            # Last punch ignored for rules; keep earliest first-punch if duplicates
            cout = None
            if out_col:
                cout_raw = (raw.get(out_col) or "").strip()
                if cout_raw and cout_raw != cin_raw:
                    cout = _parse_dt(cout_raw, on_date, tz)

            slot = punches[key].get(on_date)
            if slot is None:
                punches[key][on_date] = {"check_in": cin, "check_out": cout}
            else:
                if cin < (slot["check_in"] or cin):
                    slot["check_in"] = cin
                if cout and (slot["check_out"] is None or cout > slot["check_out"]):
                    slot["check_out"] = cout
        except Exception as exc:  # noqa: BLE001
            errors.append(ImportErrorRow(row=idx, message=str(exc)))

    if not punches:
        raise ValidationFailed(
            "No valid punch rows found. Expect columns: First Name, Date, First Punch."
        )

    all_dates = [d for emp_days in punches.values() for d in emp_days]
    period_start, period_end = _require_single_month(all_dates)
    roster_keys = sorted(punches.keys(), key=lambda k: display_names[k].lower())

    mode = (saturday_off_mode or "second_saturday").strip().lower()
    if mode not in ("second_saturday", "date"):
        raise ValidationFailed("saturday_off_mode must be second_saturday or date")

    forced_saturday_offs: set[date] = set()
    if mode == "date":
        if saturday_off_date is None:
            raise ValidationFailed("Pick a Saturday off date, or choose Recommended")
        if saturday_off_date.weekday() != 5:
            raise ValidationFailed(
                f"{saturday_off_date.isoformat()} is not a Saturday — pick a Saturday"
            )
        if not (period_start <= saturday_off_date <= period_end):
            raise ValidationFailed(
                f"Saturday off {saturday_off_date.isoformat()} is outside this file's period "
                f"({period_start} → {period_end})"
            )
        forced_saturday_offs = {saturday_off_date}
    else:
        forced_saturday_offs = _second_saturdays(period_start, period_end)

    # Classify each calendar day using Excel roster (not DB headcount)
    day_types: dict[date, str] = {}
    d = period_start
    while d <= period_end:
        if d.weekday() == 6:
            day_types[d] = "sunday_off"
        elif d in configured_holidays:
            day_types[d] = "configured_holiday"
        elif d in forced_saturday_offs:
            day_types[d] = "saturday_off"
        else:
            present_n = sum(1 for k in roster_keys if punches.get(k, {}).get(d, {}).get("check_in"))
            total_n = max(len(roster_keys), 1)
            absent_rate = 1.0 - (present_n / total_n)
            # Other weekdays only: majority empty → company / gazetted-style off
            if d.weekday() != 5 and absent_rate >= majority:
                day_types[d] = "auto_holiday"
            else:
                day_types[d] = "working"
        d += timedelta(days=1)

    imported = 0
    employee_reports: list[PeriodEmployeeReport] = []

    for key in roster_keys:
        display = display_names[key]
        emp = linked_employee.get(key)
        emp_punches = punches.get(key, {})
        late_events: list[LateEvent] = []
        half_day_dates: list[date] = []
        absent_dates: list[date] = []
        present_dates: list[date] = []
        ot_dates: list[date] = []
        sunday_dates: list[date] = []
        daily_entries: list[PeriodDayEntry] = []
        days_late = days_half = days_absent = days_present = days_sunday = 0

        d = period_start
        while d <= period_end:
            dtype = day_types[d]
            punch = emp_punches.get(d)
            cin = punch["check_in"] if punch else None
            cout = punch["check_out"] if punch else None
            local_in = cin.astimezone(tz).time() if cin else None

            status: str
            notes: str | None = None

            if dtype == "sunday_off":
                if cin is not None:
                    status = "holiday"
                    notes = "Sunday present — OT (not counted as present)"
                    ot_dates.append(d)
                    sunday_dates.append(d)
                    days_sunday += 1
                else:
                    status = "holiday"
            elif dtype in ("saturday_off", "configured_holiday", "auto_holiday"):
                if cin is not None:
                    status = "holiday"
                    notes = "OT — worked on Saturday off / company holiday"
                    ot_dates.append(d)
                else:
                    status = "holiday"
            else:
                if cin is None:
                    status = "absent"
                    days_absent += 1
                    absent_dates.append(d)
                elif local_in and local_in > half_after:
                    status = "half_day"
                    days_half += 1
                    days_late += 1
                    days_present += 1
                    present_dates.append(d)
                    half_day_dates.append(d)
                    late_events.append(
                        LateEvent(date=d, check_in_time=local_in.strftime("%H:%M"))
                    )
                elif local_in and local_in > late_after:
                    status = "late"
                    days_late += 1
                    days_present += 1
                    present_dates.append(d)
                    late_events.append(
                        LateEvent(date=d, check_in_time=local_in.strftime("%H:%M"))
                    )
                else:
                    status = "present"
                    days_present += 1
                    present_dates.append(d)

            check_in_time = local_in.strftime("%H:%M") if local_in else None
            check_out_time = None
            if cout is not None:
                check_out_time = cout.astimezone(tz).time().strftime("%H:%M")

            daily_entries.append(
                PeriodDayEntry(
                    date=d,
                    weekday=d.strftime("%A"),
                    status=status,
                    day_type=dtype,
                    check_in_time=check_in_time,
                    check_out_time=check_out_time,
                    notes=notes,
                )
            )

            if persist and emp is not None:
                existing = (
                    db.query(AttendanceRecord)
                    .filter(AttendanceRecord.employee_id == emp.id, AttendanceRecord.date == d)
                    .one_or_none()
                )
                # Always persist explicit Saturday-off / holiday so the day is never treated as absent later
                should_write = (
                    cin is not None
                    or status in ("absent", "half_day", "late", "present")
                    or dtype in ("saturday_off", "configured_holiday", "auto_holiday", "sunday_off")
                )
                if existing is None and should_write:
                    db.add(
                        AttendanceRecord(
                            employee_id=emp.id,
                            date=d,
                            check_in=cin,
                            check_out=cout,
                            source="import",
                            status=status,
                            notes=notes,
                        )
                    )
                    imported += 1
                elif existing is not None:
                    if cin is not None:
                        existing.check_in = cin
                    if cout is not None:
                        existing.check_out = cout
                    existing.source = "import"
                    existing.status = status
                    if notes:
                        existing.notes = notes
                    imported += 1

            d += timedelta(days=1)

        late_off_days = days_late // lates_per_off
        tenure_m = _tenure_months(emp.date_joined, period_end) if emp else 0
        leave_allowance = monthly_leave if emp and tenure_m >= leave_after_months else 0
        leave_used = min(leave_allowance, days_absent)
        raw_absents_after_leave = max(0, days_absent - leave_used)
        # Late absents stay in late_off_days — they are not regular absents (no-show only).
        deduction_days = (
            Decimal(raw_absents_after_leave)
            + Decimal(late_off_days)
            + (Decimal(days_half) * Decimal("0.5"))
        )
        ot_days = len(ot_dates)
        excel_id = excel_ids.get(key)
        base = Decimal(str(emp.base_salary or 0)) if emp and emp.base_salary is not None else Decimal("0")
        per_day = (base / Decimal(month_days)) if base else Decimal("0")
        estimated_deduction = (deduction_days * per_day).quantize(Decimal("0.01"))
        estimated_ot = (Decimal(ot_days) * per_day).quantize(Decimal("0.01"))
        estimated_net = (base - estimated_deduction + estimated_ot).quantize(Decimal("0.01"))

        employee_reports.append(
            PeriodEmployeeReport(
                employee_id=emp.id if emp else None,
                employee_code=emp.employee_code if emp else excel_id,
                excel_employee_id=excel_id,
                full_name=display,
                matched_employee=emp is not None,
                base_salary=base if emp and emp.base_salary is not None else None,
                tenure_months=tenure_m,
                leave_allowance=leave_allowance,
                leave_used=leave_used,
                days_present=days_present,
                days_late=days_late,
                days_half_day=days_half,
                days_sunday_present=days_sunday,
                days_absent=days_absent,
                absents_after_leave=raw_absents_after_leave,
                late_off_days=late_off_days,
                overtime_bonus_days=ot_days,
                deduction_days=float(deduction_days),
                per_day_rate=float(per_day),
                estimated_deduction_amount=float(estimated_deduction),
                estimated_overtime_amount=float(estimated_ot),
                estimated_net_salary=float(estimated_net),
                late_events=late_events,
                half_day_dates=half_day_dates,
                sunday_dates=sunday_dates,
                absent_dates=absent_dates,
                overtime_dates=ot_dates,
                daily_entries=daily_entries,
            )
        )

    unmatched: list[UnmatchedAttendancePerson] = []
    for key in roster_keys:
        if key in linked_employee:
            continue
        unmatched.append(
            UnmatchedAttendancePerson(
                full_name=display_names.get(key, key),
                excel_employee_id=excel_ids.get(key),
            )
        )
        errors.append(
            ImportErrorRow(
                row=0,
                message=(
                    f"{display_names.get(key, key)} is in the Excel file but is not an employee yet. "
                    "Add them as an employee below, then re-upload professionally to save their attendance and salary."
                ),
            )
        )

    payroll_periods: list[PayrollPeriodSaved] = []
    payroll_message: str | None = None
    if persist:
        db.flush()
        matched_ids = {emp.id for emp in linked_employee.values()}
        try:
            with db.begin_nested():
                payroll_periods, payroll_message = _persist_payroll_for_import(
                    db,
                    auth,
                    period_start=period_start,
                    period_end=period_end,
                    matched_employee_ids=matched_ids,
                )
        except Exception as exc:  # noqa: BLE001
            logger.exception("Payroll save after professional attendance import failed")
            payroll_message = (
                f"Attendance was saved for {period_start.strftime('%B %Y')}, "
                f"but payroll could not be saved: {exc}"
            )
        audit_service.log_from_auth(
            db,
            auth,
            action="attendance.period_report_imported",
            entity_type="attendance_import",
            entity_id=0,
            after_state={
                "filename": filename,
                "import_mode": mode_norm,
                "imported": imported,
                "errors": len(errors),
                "period_start": str(period_start),
                "period_end": str(period_end),
                "excel_people": len(roster_keys),
                "saturday_off_mode": mode,
                "saturday_off_dates": sorted(d.isoformat() for d in forced_saturday_offs),
                "extra_holiday_dates": sorted(d.isoformat() for d in extra_holidays),
                "payroll_periods": [
                    {
                        "period_month": p.period_month,
                        "period_year": p.period_year,
                        "employees_saved": p.employees_saved,
                    }
                    for p in payroll_periods
                ],
            },
        )

    day_classifications = [
        DayClassification(
            date=day,
            day_type=dtype,
            weekday=day.strftime("%A"),
        )
        for day, dtype in sorted(day_types.items())
        if dtype != "working"
    ]

    first_payroll = payroll_periods[0] if payroll_periods else None
    return AttendancePeriodReport(
        period_start=period_start,
        period_end=period_end,
        month_days=month_days,
        majority_absent_threshold=majority,
        late_after=late_after.strftime("%H:%M"),
        half_day_after=half_after.strftime("%H:%M"),
        lates_per_off=lates_per_off,
        imported_rows=imported,
        errors=errors,
        non_working_days=day_classifications,
        employees=employee_reports,
        unmatched_people=unmatched,
        saturday_off_mode=mode,
        saturday_off_dates=sorted(forced_saturday_offs),
        extra_holiday_dates=sorted(extra_holidays),
        import_mode=mode_norm,  # type: ignore[arg-type]
        persisted=persist,
        payroll_saved=bool(payroll_periods),
        payroll_period_month=first_payroll.period_month if first_payroll else period_start.month,
        payroll_period_year=first_payroll.period_year if first_payroll else period_start.year,
        payroll_employees_saved=sum(p.employees_saved for p in payroll_periods),
        payroll_message=payroll_message,
        payroll_periods=payroll_periods,
    )


def ensure_office_policy_config(db: Session) -> None:
    row = db.query(SystemConfig).filter_by(key="attendance.office_policy").one_or_none()
    defaults = {
        "late_after": "09:40",
        "half_day_after": "11:30",
        "majority_absent_threshold": 0.9,
        "lates_per_off": 3,
        "month_days": 30,
        "leave_after_months": 6,
        "monthly_leave_allowance": 1,
    }
    if row is None:
        db.add(SystemConfig(key="attendance.office_policy", value=defaults))
        return
    if isinstance(row.value, dict):
        merged = {**defaults, **row.value, "majority_absent_threshold": 0.9}
        row.value = merged


def create_employees_from_excel(
    db: Session, auth: AuthContext, payload: AttendanceEmployeesFromExcelCreate
) -> AttendanceEmployeesFromExcelResult:
    """Create stub employees from unmatched Excel names. Code = Excel Employee ID."""
    from app.services.seed_service import seed_default_department

    seed_default_department(db)
    dept = db.query(Department).order_by(Department.id).first()
    if dept is None:
        raise ValidationFailed("No department exists — create one before adding employees")

    created: list[UnmatchedAttendancePerson] = []
    skipped: list[str] = []
    indexes = build_employee_indexes(db)

    for person in payload.people:
        name = (person.full_name or "").strip()
        if not name:
            skipped.append("(empty name)")
            continue
        code = _normalize_excel_id(person.excel_employee_id or "")
        existing = match_employee(code=code or None, name=name, email=None, indexes=indexes)
        if existing is not None:
            skipped.append(f"{name} — already exists as {existing.employee_code}")
            continue
        if not code:
            skipped.append(f"{name} — Excel Employee ID is missing")
            continue
        taken = db.query(Employee).filter(Employee.employee_code == code).one_or_none()
        if taken is not None:
            skipped.append(f"{name} — employee ID {code} is already used by {taken.full_name}")
            continue
        emp = Employee(
            employee_code=code,
            full_name=name,
            department_id=dept.id,
            role_title=dept.name or "Employee",
            employment_type="full_time",
            status="active",
        )
        db.add(emp)
        db.flush()
        audit_service.log_from_auth(
            db,
            auth,
            action="employee.created",
            entity_type="employee",
            entity_id=emp.id,
            after_state={"employee_code": code, "full_name": name, "source": "attendance_excel"},
        )
        created.append(UnmatchedAttendancePerson(full_name=name, excel_employee_id=code))
        indexes = build_employee_indexes(db)

    return AttendanceEmployeesFromExcelResult(
        created=len(created),
        skipped=skipped,
        employees=created,
    )
